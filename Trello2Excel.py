#从Trello取信息

import requests
import json
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

# grab the active worksheet
ws1 = wb.active
ws1.title = "研发二部"
ws2 = wb.create_sheet("产品平台项目") 
ws3 = wb.create_sheet("量产和生产自动化项目") 



# === Board ID：===
# 研发二部 W8rR2s3M
# 产品平台项目 0r30QEGn
# 量产和生产自动化项目 FJXClMOs
# 100G光模块项目 858NVbBr
# RD2_bak VxdIMsn7

Board_id=['W8rR2s3M','0r30QEGn','FJXClMOs']
bi=0
for ws in wb:
    ws.append(['List','Label' , 'Task', 'Due Date', 'Member', '计划完成时间', '工时', '卡片建立时间', 'URL Link'])
    
    url = "https://api.trello.com/1/boards/" + Board_id[bi] + "/lists" 

    querystring = {"key":"f5f9a413bd62fd7a027f64da998fba16","token":"e8be1df2aea34c7213a8adc0e380bf457f8ca60a2a12f5efd75c13a9440cf518"}

    response = requests.request("GET", url, params=querystring)
    lists=response.json()
    print("lists number:",len(lists))

    url = "https://api.trello.com/1/boards/" + Board_id[bi] + "/members" 

    querystring = {"key":"f5f9a413bd62fd7a027f64da998fba16","token":"e8be1df2aea34c7213a8adc0e380bf457f8ca60a2a12f5efd75c13a9440cf518","fields":"fullName"}

    response = requests.request("GET", url, params=querystring)
    board_members=response.json()
    print('Members number:',len(board_members))


    dict_members={board_members[i]['id']:board_members[i]['fullName'] for i in range(len(board_members))}
    #print(dict_members)


    i=0
    j=0
    crow=1
    for i in range(len(lists)):
        #print(lists[i]['name'])
        #ws.cell(row=crow+1, column=1,value=(lists[i]['name']))
        url = "https://api.trello.com/1/lists/"+lists[i]['id']+"/cards"

        querystring = {"key":"f5f9a413bd62fd7a027f64da998fba16","token":"e8be1df2aea34c7213a8adc0e380bf457f8ca60a2a12f5efd75c13a9440cf518"}

        response = requests.request("GET", url, params=querystring)
        cards=response.json()
        #print('cards number:',len(cards))
        #print(cards[0])

        for j in range(len(cards)):
        #print(cards[i]['id'])
            print(cards[j]['name'])
            labels_dict=cards[j]['labels']
            #print(str(labels_dict))
            if str(labels_dict)=='[]':
                #print('[]')
                pass
            else:
                #print(labels_dict[0]['name'])
                ws.cell(row=crow+1+j, column=2,value=labels_dict[0]['name'])
            ws.cell(row=crow+1+j, column=3,value=(cards[j]['name']))
            ws.cell(row=crow+1+j, column=1,value=(lists[i]['name']))
            ws.cell(row=crow+1+j, column=8,value=datetime.fromtimestamp(int(cards[j]['id'][0:8],16)).isoformat()[:10]) #卡片建立时间
            # https://help.trello.com/article/759-getting-the-time-a-card-or-board-was-created

            if cards[j]['due']==None:
                #print('None Due date')
                ws.cell(row=crow+1+j, column=4,value=(cards[j]['due']))
            else:
                #print(cards[j]['due'][:10])
                ws.cell(row=crow+1+j, column=4,value=(cards[j]['due'][:10]))
            
            if len(cards[j]['idMembers'])>0:
                #print(cards[j]['idMembers'][0])
                members=dict_members[cards[j]['idMembers'][0]]
                #print(members)
                ws.cell(row=crow+1+j, column=5,value=members)

            if len(cards)<25:
                url = "https://api.trello.com/1/cards/"+cards[j]['id']+"/customFieldItems"
                querystring = {"key":"f5f9a413bd62fd7a027f64da998fba16","token":"e8be1df2aea34c7213a8adc0e380bf457f8ca60a2a12f5efd75c13a9440cf518"}
                response = requests.request("GET", url, params=querystring)
                if len(response.json())>0:
                    #print('customFieldItems:',response.json()[-1]['value'])
                    ws.cell(row=crow+1+j, column=6,value=response.json()[-1]['value']['date'][:10])
                if len(response.json())>1:
                    ws.cell(row=crow+1+j, column=7,value=response.json()[-2]['value']['number'])

                #url = "https://api.trello.com/1/cards/"+cards[j]['id']+"/actions?limit=5"
                #querystring = {"key":"f5f9a413bd62fd7a027f64da998fba16","token":"e8be1df2aea34c7213a8adc0e380bf457f8ca60a2a12f5efd75c13a9440cf518","filter":"all"}
                #response = requests.request("GET", url, params=querystring)
                #ws.cell(row=crow+1+j, column=8,value=response.json()[-1]['date'][:10])
                

            ws.cell(row=crow+1+j, column=9,value=cards[j]['shortUrl'])
            print('Cards number:',j)


        crow=j+crow+1
    
    bi=bi+1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30


# Save the file
wb.save("sample.xlsx")

print('Successfully Done!')

